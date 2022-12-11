import datetime
from openpyxl import load_workbook
import calendar
from datetime import datetime, timedelta

# Define a class to represent a shift
class Shift:
  # Initialize a new shift with a start and end time
  def __init__(self, start_time, end_time):
    self.start_time = start_time
    self.end_time = end_time
    
  # Create a new shift with the given start and end date and time
  def create_shift(start_date, start_time, end_date, end_time):
    # Parse the start and end date and time into datetime objects
    start_datetime = datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M")
    end_datetime = datetime.strptime(f"{end_date} {end_time}", "%Y-%m-%d %H:%M")
        
    # Return a new Shift object with the parsed start and end datetime
    return Shift(start_datetime, end_datetime)

# Define a class to represent a worker
class Worker:
  # Initialize a new worker with empty shift data
  def __init__(self):
    self.shifts = []
    
  # Add a shift to the worker's shift data
  def add_shift(self, shift):
    self.shifts.append(shift)
    
  # Calculate the total number of hours worked by the worker
  #def get_hours(self):
  #  total_hours = 0
  #  for shift in self.shifts:
  #    total_hours += shift.end_time - shift.start_time
  #    return total_hours
    
  # Get a list of the worker's shifts
  def get_shifts(self):
    return self.shifts

#reading a colored cell in excel

# Load the Excel workbook
workbook = load_workbook("./November.xlsx")

# Get the active sheet
sheet = workbook.active

year = 2022
month = 11
worker_name = "אדם".encode('utf-8')
morning_str = "בוקר".encode('utf-8')
evening_str = "ערב".encode('utf-8')
night_str = "לילה".encode('utf-8')
friday_str = "שישי".encode('utf-8')
bonusHours = 0

# Get the number of days in the month of February in the year 2021
daysThisMonth = calendar.monthrange(year, month)[1]

# Print the result
#print(daysThisMonth)  # Output: 30 # for novmber

# Create a new worker
worker = Worker()

for row in sheet.iter_rows():
  for cell in row:
    if cell.value == worker_name.decode('utf-8'):
      workers_row = row

# Scan the cells in the sheet
for row in sheet.iter_rows():
  if row == workers_row:
    for cell in row:
      if 3 == len(cell.coordinate):
        day =  ord(cell.coordinate[0]) - 64
        
      elif 4 == len(cell.coordinate) and day < daysThisMonth:
        #day = ord("Z") + ord(cell.coordinate[1])
        day = day + 1
      
      else:
        break
      
      #print(day)
      #my_datetime = "Saturday"
      # Create a datetime object
      if day == 1:
        my_datetime = datetime(year, month, day)
      elif 1 < day:
        day -= 1
        my_datetime = datetime(year, month, day)
      #except ValueError as error:
      #  print(error)
      
      #print(my_datetime)
      # Use the strftime method to convert the datetime object to a string with the desired format
      my_datetime_str = my_datetime.strftime("%A")
      #print(my_datetime_str)

      if cell.value == morning_str.decode('utf-8'):
        #print(f"Cell {cell.coordinate} is a morning shift")
        new_shift = Shift.create_shift(f"{str(year)}-{str(month)}-{str(day)}", "07:00", f"{str(year)}-{str(month)}-{str(day)}", "16:00")
        worker.add_shift(new_shift)

      elif cell.value == evening_str.decode('utf-8'):
        #print(f"Cell {cell.coordinate} is a evening shift")
        new_shift = Shift.create_shift(f"{str(year)}-{str(month)}-{str(day)}", "16:00", f"{str(year)}-{str(month)}-{str(day)}", "23:00")
        worker.add_shift(new_shift)
    
      elif cell.value == night_str.decode('utf-8'):
        #print(f"Cell {cell.coordinate} is a night shift")
        if "Saturday" == my_datetime_str:
          next_Day = day + 1
          new_shift = Shift.create_shift(f"{str(year)}-{str(month)}-{str(day)}", "22:00", f"{str(year)}-{str(month)}-{str(next_Day)}", "07:00")
          bonusHours += 2
          worker.add_shift(new_shift)
        else:
          next_Day = day + 1
          new_shift = Shift.create_shift(f"{str(year)}-{str(month)}-{str(day)}", "23:00", f"{str(year)}-{str(month)}-{str(next_Day)}", "07:00")
          bonusHours += 1
          worker.add_shift(new_shift)
        #print(my_datetime_str)
          
      elif cell.value == friday_str.decode('utf-8'):
        #print(f"Cell {cell.coordinate} is a friday shift")
        new_shift = Shift.create_shift(f"{str(year)}-{str(month)}-{str(day)}", "07:00", f"{str(year)}-{str(month)}-{str(day)}", "16:00")
        worker.add_shift(new_shift)


# Calculate the total number of hours worked
#hours = worker.get_hours()

# Print the total number of hours worked
#print(f"Total hours worked: {hours}")

# Print the list of shifts
print("Shifts:")
for shift in worker.get_shifts():
  print(f"  Start: {shift.start_time}, End: {shift.end_time}")
